import io
from openpyxl import load_workbook
from datetime import date, timedelta
from io import BytesIO

from django.test import TestCase
from django.utils import timezone

from projects.models import Project
from .models import Department, DepartmentActivityRecord, SpaceMaterialRequest, Item, Category
from .department_activity_services import (
    get_department_activity_queryset,
    require_valid_department,
    get_report_date_range,
)


class DepartmentActivityServicesTests(TestCase):
    def setUp(self):
        self.project = Project.objects.create(project_code='P1', project_name='Project 1', employer='Emp1', start_date=date.today())
        self.other_project = Project.objects.create(project_code='P2', project_name='Project 2', employer='Emp2', start_date=date.today())
        self.record1 = DepartmentActivityRecord.objects.create(
            project=self.project,
            department=Department.BUILDINGS,
            date=date(2023, 1, 1),
            location='Location A',
            activity_description='Activity A',
            contractor='Contractor A',
            unit='m2',
        )
        self.record2 = DepartmentActivityRecord.objects.create(
            project=self.project,
            department=Department.MECHANICAL,
            date=date(2023, 1, 5),
            location='Location B',
            activity_description='Activity B',
            contractor='Contractor B',
            unit='m3',
        )
        self.record3 = DepartmentActivityRecord.objects.create(
            project=self.other_project,
            department=Department.BUILDINGS,
            date=date(2023, 1, 10),
            location='Location C',
            activity_description='Activity C',
            contractor='Contractor C',
            unit='kg',
        )

    def test_get_department_activity_queryset_project_isolation(self):
        qs = get_department_activity_queryset(self.project.id, {})
        self.assertEqual(qs.count(), 2)
        self.assertIn(self.record1, qs)
        self.assertIn(self.record2, qs)
        self.assertNotIn(self.record3, qs)

    def test_get_department_activity_queryset_department_filter(self):
        qs = get_department_activity_queryset(self.project.id, {'department': Department.BUILDINGS})
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs.first(), self.record1)

    def test_get_department_activity_queryset_invalid_department(self):
        qs = get_department_activity_queryset(self.project.id, {'department': 'invalid_dept'})
        self.assertEqual(qs.count(), 0)

    def test_get_department_activity_queryset_date_filter(self):
        qs = get_department_activity_queryset(self.project.id, {'date_from': '2023-01-02'})
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs.first(), self.record2)

        qs = get_department_activity_queryset(self.project.id, {'date_to': '2023-01-02'})
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs.first(), self.record1)

        qs = get_department_activity_queryset(self.project.id, {'date_from': '2023-01-01', 'date_to': '2023-01-05'})
        self.assertEqual(qs.count(), 2)

    def test_get_department_activity_queryset_text_filters(self):
        qs = get_department_activity_queryset(self.project.id, {'location': 'Location A'})
        self.assertEqual(qs.count(), 1)

        qs = get_department_activity_queryset(self.project.id, {'activity_description': 'Activity B'})
        self.assertEqual(qs.count(), 1)

        qs = get_department_activity_queryset(self.project.id, {'contractor': 'Contractor A'})
        self.assertEqual(qs.count(), 1)

        qs = get_department_activity_queryset(self.project.id, {'unit': 'm3'})
        self.assertEqual(qs.count(), 1)

    def test_get_department_activity_queryset_search(self):
        qs = get_department_activity_queryset(self.project.id, {'search': 'Location'})
        self.assertEqual(qs.count(), 2)

        qs = get_department_activity_queryset(self.project.id, {'search': 'Activity A'})
        self.assertEqual(qs.count(), 1)

    def test_get_department_activity_queryset_ordering(self):
        qs = get_department_activity_queryset(self.project.id, {'ordering': 'date'})
        self.assertEqual(qs.first(), self.record1)

        qs = get_department_activity_queryset(self.project.id, {'ordering': '-date'})
        self.assertEqual(qs.first(), self.record2)

        # Default ordering
        qs = get_department_activity_queryset(self.project.id, {})
        self.assertEqual(qs.first(), self.record2)

    def test_require_valid_department(self):
        self.assertEqual(require_valid_department(Department.BUILDINGS), Department.BUILDINGS)
        self.assertIsNone(require_valid_department('invalid'))
        self.assertIsNone(require_valid_department(None))

    def test_get_report_date_range(self):
        today = timezone.localdate()
        yesterday = today - timedelta(days=1)
        start_weekly = today - timedelta(days=6)

        d1, d2 = get_report_date_range('daily')
        self.assertEqual(d1, yesterday)
        self.assertEqual(d2, yesterday)

        w1, w2 = get_report_date_range('weekly')
        self.assertEqual(w1, start_weekly)
        self.assertEqual(w2, today)

from .department_activity_io import (
    export_activities_to_xlsx,
    _map_header_row,
    _parse_date_value,
    import_activities_from_xlsx,
    generate_activity_report_pdf,
    department_display_label,
)
from openpyxl import Workbook
import datetime

class DepartmentActivityIOTests(TestCase):
    def setUp(self):
        self.project = Project.objects.create(project_code='P1', project_name='Project 1', employer='Emp1', start_date=date.today())
        self.record1 = DepartmentActivityRecord.objects.create(
            project=self.project,
            department=Department.BUILDINGS,
            date=date(2023, 1, 1),
            location='Location A',
            activity_description='Activity A',
            contractor='Contractor A',
            unit='m2',
            description='Desc A'
        )

    def test_export_activities_to_xlsx(self):
        records = [self.record1]
        xlsx_bytes = export_activities_to_xlsx(records)
        self.assertIsInstance(xlsx_bytes, bytes)
        self.assertTrue(len(xlsx_bytes) > 0)

        # We can try to load it to make sure it's valid
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb.active
        self.assertEqual(ws.title, 'Activity log')
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], ('date', 'location', 'activity_description', 'contractor', 'unit', 'description'))
        self.assertEqual(rows[1], ('2023-01-01', 'Location A', 'Activity A', 'Contractor A', 'm2', 'Desc A'))

    def test_map_header_row(self):
        header_row = ('تاریخ', 'موقعیت', 'Activity Description', 'پیمانکار', 'Unit', 'توضیحات', 'Unknown')
        col_map = _map_header_row(header_row)
        self.assertEqual(col_map, {
            0: 'date',
            1: 'location',
            2: 'activity_description',
            3: 'contractor',
            4: 'unit',
            5: 'description'
        })

    def test_parse_date_value(self):
        self.assertEqual(_parse_date_value(date(2023, 1, 1)), date(2023, 1, 1))
        self.assertEqual(_parse_date_value(datetime.datetime(2023, 1, 1, 12, 0)), date(2023, 1, 1))
        self.assertEqual(_parse_date_value('2023-01-01'), date(2023, 1, 1))
        self.assertEqual(_parse_date_value('2023-01-01T12:00:00'), date(2023, 1, 1))
        self.assertIsNone(_parse_date_value(None))
        self.assertIsNone(_parse_date_value(''))
        self.assertIsNone(_parse_date_value('invalid-date'))

    def _create_xlsx_bytes(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_import_activities_from_xlsx_success(self):
        rows = [
            ('date', 'location', 'activity_description', 'contractor', 'unit', 'description'),
            ('2023-01-02', 'Loc B', 'Act B', 'Con B', 'kg', 'Desc B'),
            ('2023-01-03', 'Loc C', 'Act C', 'Con C', 'm', '')
        ]
        xlsx_bytes = self._create_xlsx_bytes(rows)

        created, errors = import_activities_from_xlsx(self.project, Department.BUILDINGS, xlsx_bytes)
        self.assertEqual(created, 2)
        self.assertEqual(len(errors), 0)

        self.assertEqual(DepartmentActivityRecord.objects.filter(project=self.project, date=date(2023, 1, 2)).count(), 1)
        self.assertEqual(DepartmentActivityRecord.objects.filter(project=self.project, date=date(2023, 1, 3)).count(), 1)

    def test_import_activities_from_xlsx_missing_headers(self):
        rows = [
            ('date', 'location', 'activity_description', 'contractor'), # Missing unit
            ('2023-01-02', 'Loc B', 'Act B', 'Con B')
        ]
        xlsx_bytes = self._create_xlsx_bytes(rows)

        created, errors = import_activities_from_xlsx(self.project, Department.BUILDINGS, xlsx_bytes)
        self.assertEqual(created, 0)
        self.assertEqual(len(errors), 1)
        self.assertIn('_sheet', errors[0]['errors'])

    def test_import_activities_from_xlsx_validation_errors(self):
        rows = [
            ('date', 'location', 'activity_description', 'contractor', 'unit', 'description'),
            ('', 'Loc B', 'Act B', 'Con B', 'kg', ''), # Missing date
            ('2023-01-03', '', 'Act C', 'Con C', 'm', ''), # Missing location
        ]
        xlsx_bytes = self._create_xlsx_bytes(rows)

        created, errors = import_activities_from_xlsx(self.project, Department.BUILDINGS, xlsx_bytes)
        self.assertEqual(created, 0)
        self.assertEqual(len(errors), 2)
        self.assertIn('date', errors[0]['errors'])
        self.assertIn('location', errors[1]['errors'])

    def test_generate_activity_report_pdf(self):
        records = [self.record1]
        pdf_bytes = generate_activity_report_pdf(
            business=self.project,
            department=Department.BUILDINGS,
            department_label=department_display_label(Department.BUILDINGS),
            period_label='Daily report',
            date_from=date(2023, 1, 1),
            date_to=date(2023, 1, 1),
            records=records
        )
        self.assertIsInstance(pdf_bytes, bytes)
        self.assertTrue(len(pdf_bytes) > 0)
        self.assertTrue(pdf_bytes.startswith(b'%PDF-'))

        pdf_bytes_empty = generate_activity_report_pdf(
            business=self.project,
            department=Department.BUILDINGS,
            department_label=department_display_label(Department.BUILDINGS),
            period_label='Daily report',
            date_from=date(2023, 1, 1),
            date_to=date(2023, 1, 1),
            records=[]
        )
        self.assertIsInstance(pdf_bytes_empty, bytes)
        self.assertTrue(len(pdf_bytes_empty) > 0)
        self.assertTrue(pdf_bytes_empty.startswith(b'%PDF-'))

    def test_department_display_label(self):
        self.assertEqual(department_display_label(Department.BUILDINGS), 'ابنیه')
        self.assertEqual(department_display_label('unknown'), 'unknown')
from django.urls import reverse
from rest_framework.test import APIClient
from authentication.models import User
from master_data.models import ProjectMember, Role
from permissions.constants import DEFAULT_ROLE_PERMISSIONS

class DepartmentActivityDataViewsTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='testuser', password='password', mobile='09123456789')
        self.project = Project.objects.create(project_code='P1', project_name='Project 1', employer='Emp1', start_date=date.today())

        from master_data.models import ProjectMemberRole
        self.role, _ = Role.objects.get_or_create(role_name='project_manager')
        self.member = ProjectMember.objects.create(project=self.project, user=self.user)
        ProjectMemberRole.objects.create(member=self.member, role=self.role)
        self.user.is_superuser = True
        self.user.save()

        self.record1 = DepartmentActivityRecord.objects.create(
            project=self.project,
            department=Department.BUILDINGS,
            date=date.today(),
            location='Location A',
            activity_description='Activity A',
            contractor='Contractor A',
            unit='m2',
            description='Desc A'
        )

        self.client.force_authenticate(user=self.user)

    def test_export_view(self):
        url = reverse('project-department-activity-record-export', kwargs={'project_pk': self.project.id})

        # Missing department
        response = self.client.get(url)
        self.assertEqual(response.status_code, 400)

        import uuid
        # Invalid project -> returns 404 because `_get_business` doesn't find it
        invalid_url = reverse('project-department-activity-record-export', kwargs={'project_pk': uuid.uuid4()})
        response = self.client.get(invalid_url, {'department': Department.BUILDINGS})
        self.assertEqual(response.status_code, 404)

        # Valid request
        response = self.client.get(url, {'department': Department.BUILDINGS})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_import_view(self):
        url = reverse('project-department-activity-record-import', kwargs={'project_pk': self.project.id})

        # Valid xlsx bytes from io tests
        wb = Workbook()
        ws = wb.active
        ws.append(('date', 'location', 'activity_description', 'contractor', 'unit', 'description'))
        ws.append(('2023-01-02', 'Loc B', 'Act B', 'Con B', 'kg', 'Desc B'))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import uuid
        # Invalid project -> returns 404 because `_get_business` doesn't find it
        invalid_url = reverse('project-department-activity-record-import', kwargs={'project_pk': uuid.uuid4()})
        response = self.client.post(invalid_url, {'department': Department.BUILDINGS, 'file': buf}, format='multipart')
        self.assertEqual(response.status_code, 404)

        buf.seek(0)

        from django.core.files.uploadedfile import SimpleUploadedFile
        # Valid request
        # department must be passed in the URL for this view
        file_obj = SimpleUploadedFile("test.xlsx", buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url_with_param = url + f'?department={Department.BUILDINGS}'
        response = self.client.post(url_with_param, {'file': file_obj}, format='multipart')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], 1)

    def test_daily_report_view(self):
        url = reverse('project-department-activity-record-daily-report', kwargs={'project_pk': self.project.id})

        # Missing department
        response = self.client.get(url)
        self.assertEqual(response.status_code, 400)

        # Valid request
        response = self.client.get(url, {'department': Department.BUILDINGS})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_weekly_report_view(self):
        url = reverse('project-department-activity-record-weekly-report', kwargs={'project_pk': self.project.id})

        # Missing department
        response = self.client.get(url)
        self.assertEqual(response.status_code, 400)

        # Valid request
        response = self.client.get(url, {'department': Department.BUILDINGS})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

class ViewsTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='testuser_views', password='password', mobile='09123456788')
        self.user.is_superuser = True
        self.user.save()

        self.project = Project.objects.create(project_code='P3', project_name='Project 3', employer='Emp1', start_date=date.today())

        from master_data.models import ProjectMemberRole
        self.role, _ = Role.objects.get_or_create(role_name='project_manager')
        self.member = ProjectMember.objects.create(project=self.project, user=self.user)
        ProjectMemberRole.objects.create(member=self.member, role=self.role)

        self.client.force_authenticate(user=self.user)

        self.smr = SpaceMaterialRequest.objects.create(
            project=self.project,
            block_number=1,
            floor_number=2,
            unit_number='12B',
            space_name='Living Room',
            material_code='MAT01',
            item_description='Ceramic Tiles'
        )

        self.dar = DepartmentActivityRecord.objects.create(
            project=self.project,
            department=Department.ELECTRICAL,
            date=date.today(),
            location='Floor 1',
            activity_description='Wiring',
            contractor='ElecCorp',
            unit='m'
        )

    def test_space_material_request_list(self):
        url = reverse('project-space-material-request-list', kwargs={'project_pk': self.project.id})

        # Test basic list
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['results']), 1)
        self.assertEqual(response.data['results'][0]['id'], self.smr.id)

        # Test filters
        response = self.client.get(url, {'block_number': 1})
        self.assertEqual(len(response.data['results']), 1)

        response = self.client.get(url, {'block_number': 2})
        self.assertEqual(len(response.data['results']), 0)

        response = self.client.get(url, {'space_name': 'Living'})
        self.assertEqual(len(response.data['results']), 1)

    def test_space_material_request_create(self):
        url = reverse('project-space-material-request-list', kwargs={'project_pk': self.project.id})
        data = {
            'block_number': 2,
            'floor_number': 3,
            'unit_number': '3C',
            'space_name': 'Kitchen',
            'material_code': 'MAT02',
            'item_description': 'Paint',
            'unit': 'liters'
        }
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 201)
        self.assertEqual(SpaceMaterialRequest.objects.filter(project=self.project).count(), 2)

    def test_department_activity_record_list(self):
        url = reverse('project-department-activity-record-list', kwargs={'project_pk': self.project.id})

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['results']), 1)
        self.assertEqual(response.data['results'][0]['id'], self.dar.id)

        # Test filters via viewset (uses same logic as services)
        response = self.client.get(url, {'department': Department.ELECTRICAL})
        self.assertEqual(len(response.data['results']), 1)

    def test_department_activity_record_create(self):
        url = reverse('project-department-activity-record-list', kwargs={'project_pk': self.project.id})
        data = {
            'department': Department.MECHANICAL,
            'date': '2023-02-01',
            'location': 'Basement',
            'activity_description': 'Plumbing',
            'contractor': 'PlumbCo',
            'unit': 'pipes'
        }
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 201)
        self.assertEqual(DepartmentActivityRecord.objects.filter(project=self.project).count(), 2)
