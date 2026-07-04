"""
Excel export/import and PDF reports for department activity records.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any

from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from business_meta.models import Business

from .models import Department, DepartmentActivityRecord

# Global variable EXPORT_HEADERS
EXPORT_HEADERS = [
    'date',
    'location',
    'activity_description',
    'contractor',
    'unit',
    'description',
]

HEADER_ALIASES: dict[str, list[str]] = {
    'date': ['date', 'تاریخ', 'Date'],
    'location': ['location', 'موقعیت', 'Location'],
    'activity_description': [
        'activity_description',
        'activity description',
        'شرح فعالیت',
        'Activity',
    ],
    'contractor': ['contractor', 'پیمانکار', 'پیمانگار', 'Contractor'],
    'unit': ['unit', 'واحد', 'Unit'],
    'description': ['description', 'توضیحات', 'Notes', 'notes'],
}


# Function to handle export activities to xlsx
def export_activities_to_xlsx(records: list[DepartmentActivityRecord]) -> bytes:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title='Activity log')
    ws.append(EXPORT_HEADERS)
    for record in records:
        ws.append(
            [
                record.date.isoformat() if record.date else '',
                record.location,
                record.activity_description,
                record.contractor,
                record.unit,
                record.description or '',
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Function to handle  map header row
def _map_header_row(header_row: tuple[Any, ...]) -> dict[int, str]:
    alias_to_field: dict[str, str] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_to_field[alias.strip().lower()] = field

    col_to_field: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = alias_to_field.get(key)
        if field:
            col_to_field[idx] = field
    return col_to_field


# Function to handle  parse date value
def _parse_date_value(raw: Any) -> date | None:
    if raw is None:
        return None
    if hasattr(raw, 'date') and callable(getattr(raw, 'date', None)):
        try:
            return raw.date()
        except Exception:
            pass
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    parsed = parse_date(text[:10])
    if parsed:
        return parsed
    return None


# Function to handle import activities from xlsx
def import_activities_from_xlsx(
    business: Business,
    department: str,
    file_bytes: bytes,
) -> tuple[int, list[dict]]:
    created = 0
    errors: list[dict] = []

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return 0, [{'row': 0, 'errors': {'_sheet': 'No sheet in workbook.'}}]

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return 0, [{'row': 0, 'errors': {'_sheet': 'Empty sheet.'}}]

    col_to_field = _map_header_row(header_row)
    required = {'date', 'location', 'activity_description', 'contractor', 'unit'}
    if not required.issubset(set(col_to_field.values())):
        return 0, [
            {
                'row': 0,
                'errors': {
                    '_sheet': (
                        'Missing required columns. Expected headers: '
                        + ', '.join(EXPORT_HEADERS)
                    ),
                },
            }
        ]

    for row_index, row_tuple in enumerate(rows_iter, start=2):
        if row_tuple is None or all(
            cell is None or str(cell).strip() == '' for cell in row_tuple
        ):
            continue

        values: dict[str, Any] = {}
        for idx, field in col_to_field.items():
            if idx < len(row_tuple):
                raw = row_tuple[idx]
                if raw is not None and str(raw).strip() != '':
                    values[field] = raw

        row_errors: dict[str, str] = {}
        activity_date = _parse_date_value(values.get('date'))
        if not activity_date:
            row_errors['date'] = 'Valid date is required (YYYY-MM-DD).'

        for field in ('location', 'activity_description', 'contractor', 'unit'):
            val = values.get(field)
            if val is None or str(val).strip() == '':
                row_errors[field] = 'This field is required.'

        if row_errors:
            errors.append({'row': row_index, 'errors': row_errors})
            continue

        description = values.get('description')
        description_text = '' if description is None else str(description).strip()

        try:
            DepartmentActivityRecord.objects.create(
                business=business,
                department=department,
                date=activity_date,
                location=str(values['location']).strip()[:255],
                activity_description=str(values['activity_description']).strip()[:500],
                contractor=str(values['contractor']).strip()[:255],
                unit=str(values['unit']).strip()[:64],
                description=description_text,
            )
            created += 1
        except ValidationError as exc:
            errors.append({'row': row_index, 'errors': exc.message_dict})

    wb.close()
    return created, errors


# Function to handle generate activity report pdf
def generate_activity_report_pdf(
    *,
    business: Business,
    department: str,
    department_label: str,
    period_label: str,
    date_from: date,
    date_to: date,
    records: list[DepartmentActivityRecord],
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=28,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(
            f'<b>{business.name}</b> — {department_label}',
            styles['Title'],
        ),
        Paragraph(
            f'{period_label}: {date_from.isoformat()} — {date_to.isoformat()}',
            styles['Heading3'],
        ),
        Spacer(1, 12),
    ]

    if not records:
        story.append(Paragraph('No activity records in this period.', styles['Normal']))
    else:
        table_data = [
            [
                'Date',
                'Location',
                'Activity',
                'Contractor',
                'Unit',
                'Description',
            ]
        ]
        for record in records:
            table_data.append(
                [
                    record.date.isoformat(),
                    record.location,
                    record.activity_description,
                    record.contractor,
                    record.unit,
                    (record.description or '')[:200],
                ]
            )

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[70, 90, 140, 90, 50, 160],
        )
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8eef5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    return buf.getvalue()


# Function to handle department display label
def department_display_label(department: str) -> str:
    for value, label in Department.choices:
        if value == department:
            return label
    return department
