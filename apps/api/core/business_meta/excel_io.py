"""
Excel export/import for dynamic table rows. Uses openpyxl for .xlsx.
"""
import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl import load_workbook

from .models import TableDefinition, DynamicTableRow
from .services import validate_row_data

logger = logging.getLogger(__name__)


# Function to handle export table to xlsx
def export_table_to_xlsx(table: TableDefinition) -> bytes:
    """
    Export all rows of a dynamic table to .xlsx.
    First row = headers (field names); columns in field order. id, created_at, updated_at included.
    """
    field_defs = list(table.fields.all().order_by("ordering", "name"))
    headers = ["id"] + [f.name for f in field_defs] + ["created_at", "updated_at"]
    slug_order = [f.slug for f in field_defs]

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=table.name[:31])  # Excel sheet name limit
    ws.append(headers)

    for row in DynamicTableRow.objects.filter(table=table).iterator():
        cells = [str(row.pk)]
        for slug in slug_order:
            val = row.data.get(slug)
            cells.append(_cell_value(val))
        cells.append(row.created_at.isoformat().replace("+00:00", "Z") if row.created_at else "")
        cells.append(row.updated_at.isoformat().replace("+00:00", "Z") if row.updated_at else "")
        ws.append(cells)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Function to handle  cell value
def _cell_value(val: Any) -> Any:
    """Normalise value for Excel (avoid None, keep numbers/dates as-is for display)."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "true" if val else "false"
    return val


# Function to handle import rows from xlsx
def import_rows_from_xlsx(table: TableDefinition, file_bytes: bytes) -> tuple[int, list[dict]]:
    """
    Parse .xlsx and create rows. First row = headers (must match field names).
    Returns (created_count, list of per-row errors { "row": 1-based, "errors": {...} }).
    """
    field_defs = list(table.fields.all().order_by("ordering", "name"))
    name_to_slug = {f.name: f.slug for f in field_defs}
    created = 0
    errors: list[dict] = []

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return 0, [{"row": 0, "errors": {"_sheet": "No sheet in workbook."}}]

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return 0, [{"row": 0, "errors": {"_sheet": "Empty sheet."}}]

    # Map column index -> field slug (skip id, created_at, updated_at if present)
    col_to_slug: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name and name in name_to_slug:
            col_to_slug[idx] = name_to_slug[name]
        # Allow optional columns id, created_at, updated_at (we ignore them on import)

    if not col_to_slug:
        return 0, [{"row": 0, "errors": {"_sheet": "No columns matched table fields."}}]

    for row_index, row_tuple in enumerate(rows_iter, start=2):  # 2 = first data row
        if row_tuple is None:
            continue
        data: dict[str, Any] = {}
        for idx, slug in col_to_slug.items():
            if idx < len(row_tuple):
                raw = row_tuple[idx]
                if raw is not None and str(raw).strip() != "":
                    data[slug] = _normalise_import_value(raw)

        cleaned, errs = validate_row_data(field_defs, data)
        if errs:
            errors.append({"row": row_index, "errors": errs})
            continue
        DynamicTableRow.objects.create(table=table, data=cleaned)
        created += 1

    wb.close()
    return created, errors


# Function to handle  normalise import value
def _normalise_import_value(raw: Any) -> Any:
    """Convert Excel cell value to something validate_row_data accepts."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and raw == int(raw):
            return int(raw)
        return raw
    s = str(raw).strip()
    if not s:
        return None
    if s.lower() in ("true", "1", "yes"):
        return True
    if s.lower() in ("false", "0", "no"):
        return False
    return s
